#!/usr/bin/env python3
import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


def sanitize_filename(name: str) -> str:
    bad = '<>:"/\\|?*'
    out = []
    for ch in name:
        out.append("_" if ch in bad else ch)
    return "".join(out).strip() or "sheet"


def main() -> None:
    parser = argparse.ArgumentParser(description="Export each worksheet in an .xlsx file to CSV.")
    parser.add_argument(
        "--xlsx",
        default="日报及月报发送记录.xlsx",
        help="Path to source .xlsx file",
    )
    parser.add_argument(
        "--out-dir",
        default="csv_exports",
        help="Output directory for CSV files",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx).expanduser().resolve()
    out_dir = Path(args.out_dir).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(xlsx_path, data_only=True)
    for ws in wb.worksheets:
        csv_name = f"{sanitize_filename(ws.title)}.csv"
        csv_path = out_dir / csv_name
        with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if v is None else v for v in row])
        print(csv_path)


if __name__ == "__main__":
    main()
