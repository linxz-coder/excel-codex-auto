#!/usr/bin/env python3
import argparse
import os
import re
import shutil
import sys
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from export_csv_from_excel import export_workbook_to_csvs


YEARLY_DAILY_SHEET_FORMAT = "{year}年日报"
SUMMARY_TEMPLATE = """{summary_date}中心业绩日报发送情况统计：

【未发】
{missed_block}

【迟发】
{late_block}"""


def ensure_utf8_locale():
    if sys.getfilesystemencoding().lower() != "ascii":
        return
    if os.environ.get("CODEX_UTF8_REEXEC") == "1":
        return
    env = os.environ.copy()
    env["LC_ALL"] = env.get("LC_ALL") or "zh_CN.utf8"
    env["LANG"] = env.get("LANG") or "zh_CN.utf8"
    env["PYTHONIOENCODING"] = env.get("PYTHONIOENCODING") or "utf-8"
    env["CODEX_UTF8_REEXEC"] = "1"
    os.execvpe(sys.executable, [sys.executable] + sys.argv, env)


def split_centers(raw: str):
    if raw is None:
        return []
    text = raw.strip()
    if not text or text == "无":
        return []
    parts = re.split(r"[;,，；\n]+", text)
    return [part.strip() for part in parts if part.strip() and part.strip() != "无"]


def parse_counts(text):
    missed_count = 0
    late_count = 0
    if text:
        text = str(text)
        match = re.search(r"累计未发(\d+)次", text)
        if match:
            missed_count = int(match.group(1))
        match = re.search(r"迟发(\d+)次", text)
        if match:
            late_count = int(match.group(1))
    return missed_count, late_count


def format_counts(missed_count, late_count):
    parts = []
    if missed_count:
        parts.append("累计未发%d次" % missed_count)
    if late_count:
        parts.append("迟发%d次" % late_count)
    return "，".join(parts) if parts else None


def resolve_center_name(center_name, known_names):
    if center_name in known_names:
        return center_name

    exact_suffix_matches = [name for name in known_names if name.endswith(center_name)]
    if len(exact_suffix_matches) == 1:
        return exact_suffix_matches[0]

    contains_matches = [name for name in known_names if center_name in name]
    if len(contains_matches) == 1:
        return contains_matches[0]

    if not exact_suffix_matches and not contains_matches:
        raise ValueError("未找到中心：%s" % center_name)

    raise ValueError("中心名称不唯一：%s -> %s" % (center_name, ", ".join(sorted(contains_matches or exact_suffix_matches))))


def prepare_paths(args, base_dir):
    xlsx_path = make_absolute_path(args.xlsx, base_dir)
    out_dir = make_absolute_path(args.out_dir, base_dir)
    backup_dir = make_absolute_path(args.backup_dir, base_dir)

    if args.sandbox_dir:
        sandbox_dir = make_absolute_path(args.sandbox_dir, base_dir)
        sandbox_dir.mkdir(parents=True, exist_ok=True)
        sandbox_xlsx = sandbox_dir / xlsx_path.name
        shutil.copy2(str(xlsx_path), str(sandbox_xlsx))
        xlsx_path = sandbox_xlsx

        if args.out_dir == "csv_exports":
            out_dir = sandbox_dir / "csv_exports"

        if args.backup_dir == "csv_backups":
            backup_dir = sandbox_dir / "csv_backups"

    out_dir.mkdir(parents=True, exist_ok=True)
    backup_dir.mkdir(parents=True, exist_ok=True)
    return xlsx_path, out_dir, backup_dir


def make_absolute_path(raw_path, base_dir):
    path = Path(raw_path).expanduser()
    if path.is_absolute():
        return path
    return (base_dir / path).resolve()


def update_workbook(xlsx_path, sheet_name, report_date, missed_names, late_names):
    wb = load_workbook(str(xlsx_path))
    if sheet_name not in wb.sheetnames:
        raise ValueError("未找到工作表：%s" % sheet_name)
    ws = wb[sheet_name]
    ws.cell(1, 2).value = report_date

    known_names = [row[1].value for row in ws.iter_rows(min_row=2) if row[1].value]
    resolved_missed = [resolve_center_name(name, known_names) for name in missed_names]
    resolved_late = [resolve_center_name(name, known_names) for name in late_names]

    pending = {}
    for name in resolved_missed:
        pending.setdefault(name, {"missed": 0, "late": 0})
        pending[name]["missed"] += 1
    for name in resolved_late:
        pending.setdefault(name, {"missed": 0, "late": 0})
        pending[name]["late"] += 1

    updated = {}
    for row in ws.iter_rows(min_row=2):
        name = row[1].value
        if name not in pending:
            continue
        missed_count, late_count = parse_counts(row[3].value)
        missed_count += pending[name]["missed"]
        late_count += pending[name]["late"]
        row[3].value = format_counts(missed_count, late_count)
        updated[name] = row[3].value

    missing = sorted(set(pending) - set(updated))
    if missing:
        raise ValueError("这些中心未能更新：%s" % ", ".join(missing))

    wb.save(str(xlsx_path))
    return updated, resolved_missed, resolved_late


def build_summary(summary_date, resolved_missed, resolved_late, updated_counts):
    missed_lines = [name + "（" + updated_counts[name] + "）" for name in resolved_missed] or ["无"]
    late_lines = [name + "（" + updated_counts[name] + "）" for name in resolved_late] or ["无"]
    return SUMMARY_TEMPLATE.format(
        summary_date=summary_date.strftime("%Y年%-m月%-d日"),
        missed_block="\n".join(missed_lines),
        late_block="\n".join(late_lines),
    )


def main():
    ensure_utf8_locale()
    base_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description="Automate the daily report workbook update/export/backup flow.")
    parser.add_argument("--xlsx", default="日报及月报发送记录.xlsx", help="Path to the source workbook")
    parser.add_argument("--out-dir", default="csv_exports", help="Directory for exported CSV files")
    parser.add_argument("--backup-dir", default="csv_backups", help="Directory for daily backup CSV files")
    parser.add_argument("--sandbox-dir", help="Optional test directory. When set, the workbook is copied here before any changes.")
    parser.add_argument("--sheet", help="Workbook sheet name. Defaults to '<year>年日报' based on --report-date.")
    parser.add_argument("--report-date", default=datetime.now().strftime("%Y-%m-%d"), help="Workbook update date in YYYY-MM-DD format")
    parser.add_argument("--missed", default="", help="Missed centers, separated by commas/Chinese commas/semicolons")
    parser.add_argument("--late", default="", help="Late centers, separated by commas/Chinese commas/semicolons")
    args = parser.parse_args()

    report_date = datetime.strptime(args.report_date, "%Y-%m-%d")
    summary_date = report_date - timedelta(days=1)
    sheet_name = args.sheet or YEARLY_DAILY_SHEET_FORMAT.format(year=report_date.year)
    missed_names = split_centers(args.missed)
    late_names = split_centers(args.late)

    xlsx_path, out_dir, backup_dir = prepare_paths(args, base_dir)
    updated_counts, resolved_missed, resolved_late = update_workbook(
        xlsx_path=xlsx_path,
        sheet_name=sheet_name,
        report_date=report_date,
        missed_names=missed_names,
        late_names=late_names,
    )

    export_workbook_to_csvs(xlsx_path, out_dir)

    backup_name = "%s_%s.csv" % (report_date.strftime("%Y-%m-%d"), sheet_name)
    backup_path = backup_dir / backup_name
    shutil.copy2(str(out_dir / (sheet_name + ".csv")), str(backup_path))

    print("workbook:", xlsx_path)
    print("csv:", out_dir / (sheet_name + ".csv"))
    print("backup:", backup_path)
    print()
    print(build_summary(summary_date, resolved_missed, resolved_late, updated_counts))


if __name__ == "__main__":
    main()
